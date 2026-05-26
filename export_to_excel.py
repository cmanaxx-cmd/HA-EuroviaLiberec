#!/usr/bin/env python3
"""Export Modbus measurements to Excel file."""
import sqlite3
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Please install openpyxl: pip3 install openpyxl")
    exit(1)

DB_PATH = "/config/eurovia_db.sqlite"


def export_to_excel():
    """Export database to Excel."""
    try:
        # Connect to database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Get all measurements
        cursor.execute(
            """
            SELECT timestamp,
                   temp_rad1, temp_rad2, temp_rad3,
                   temp_aku1h, temp_aku1s, temp_aku2h, temp_aku2s,
                   temp_tc1, temp_tc2,
                   other_10, other_11, other_12, other_13, other_14, other_15,
                   other_16, other_17, other_18, other_19, other_20, other_21,
                   other_22, other_23, other_24, other_25
            FROM measurements
            ORDER BY timestamp DESC
            """
        )

        rows = cursor.fetchall()
        conn.close()

        if not rows:
            print("No data to export!")
            return

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Measurements"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = [
            "Čas",
            "Datum",
            "Teplota Radiátor 1",
            "Teplota Radiátor 2",
            "Teplota Radiátor 3",
            "Teplota Akumulátor 1 Horní",
            "Teplota Akumulátor 1 Spodní",
            "Teplota Akumulátor 2 Horní",
            "Teplota Akumulátor 2 Spodní",
            "Teplota TC1",
            "Teplota TC2",
            "Měřák 10",
            "Měřák 11",
            "Měřák 12",
            "Měřák 13",
            "Měřák 14",
            "Měřák 15",
            "Měřák 16",
            "Měřák 17",
            "Měřák 18",
            "Měřák 19",
            "Měřák 20",
            "Měřák 21",
            "Měřák 22",
            "Měřák 23",
            "Měřák 24",
            "Měřák 25",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            timestamp_str = row_data[0]
            try:
                dt = datetime.fromisoformat(timestamp_str)
                time_str = dt.strftime("%H:%M")
                date_str = dt.strftime("%d.%m.%Y")
            except:
                time_str = "N/A"
                date_str = "N/A"

            ws.cell(row=row_idx, column=1, value=time_str)
            ws.cell(row=row_idx, column=2, value=date_str)

            for col_idx, value in enumerate(row_data[1:], 3):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # Save file
        timestamp = datetime.now().strftime("%Y-%m-%d")
        output_path = f"/config/eurovia_export_{timestamp}.xlsx"
        wb.save(output_path)

        print(f"✅ Export completed: {output_path}")
        print(f"📊 Exported {len(rows)} rows")

    except Exception as e:
        print(f"❌ Export error: {e}")


if __name__ == "__main__":
    export_to_excel()
